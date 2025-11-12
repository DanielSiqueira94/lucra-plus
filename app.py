import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode

# ----------------------------
# CONFIGURAÇÃO DO APP
# ----------------------------
st.set_page_config(
    page_title="Lucra+ | Controle de Margem e Lucro",
    page_icon="💰",
    layout="wide"
)

# ----------------------------
# NOVO LOGIN POR USUÁRIO
# ----------------------------
usuarios = {
    "daniel": {"senha": "senha123", "plano": "Premium"},
    "mylena": {"senha": "senha456", "plano": "Free"}
}

if "user" not in st.session_state:
    st.session_state.user = None

if st.session_state.user is None:
    st.title("🔐 Login")
    usuario_input = st.text_input("Usuário:")
    senha_input = st.text_input("Senha:", type="password")

    if st.button("Entrar"):
        if usuario_input in usuarios and senha_input == usuarios[usuario_input]["senha"]:
            st.session_state.user = usuario_input
            st.rerun()
        else:
            st.error("Usuário ou senha incorretos.")
    st.stop()

# ----------------------------
# MOSTRAR PLANO NA SIDEBAR + BOTÃO SAIR
# ----------------------------
plano_usuario = usuarios[st.session_state.user]["plano"]
st.sidebar.info(f"👤 Usuário: **{st.session_state.user}**\n🏷️ Plano: **{plano_usuario}**")

if st.sidebar.button("🚪 Sair"):
    st.session_state.user = None
    st.rerun()


# ----------------------------
# FUNÇÃO DE CÁLCULO
# ----------------------------
def calcular_resultados(df_input, margem_desejada, custos_fixos, incluir_fixos=False):
    df = df_input.copy()
    df = df.rename(columns={
        "Taxa_pct": "Taxa (%)",
        "OutrosCustos": "Outros Custos (R$)"
    })

    for col in ["Taxa (%)", "Outros Custos (R$)"]:
        if col not in df.columns:
            df[col] = 0.0

    df["Custo"] = pd.to_numeric(df["Custo"], errors="coerce").fillna(0.0)
    df["Preco"] = pd.to_numeric(df["Preco"], errors="coerce").fillna(0.0)
    df["Taxa (%)"] = pd.to_numeric(df["Taxa (%)"], errors="coerce").fillna(0.0)
    df["Outros Custos (R$)"] = pd.to_numeric(df["Outros Custos (R$)"], errors="coerce").fillna(0.0)

    t = df["Taxa (%)"] / 100
    m = margem_desejada / 100

    df["Taxa (R$)"] = df["Preco"] * t
    df["Lucro Líquido (R$)"] = df["Preco"] - df["Custo"] - df["Taxa (R$)"] - df["Outros Custos (R$)"]
    df["Margem Atual (%)"] = np.where(df["Preco"] != 0, (df["Lucro Líquido (R$)"] / df["Preco"]) * 100, 0)
    df["Margem Desejada (%)"] = margem_desejada
    df["Preço Ideal (R$)"] = np.where(
        1 - t - m > 0, 
        (df["Custo"] + df["Outros Custos (R$)"]) / (1 - t - m), 
        np.nan
    )
    df["Diferença Preço Ideal (%)"] = np.where(
        df["Preco"] != 0, 
        ((df["Preço Ideal (R$)"] - df["Preco"]) / df["Preco"]) * 100, 
        np.nan
    )
    df["Ponto de Equilíbrio (unid)"] = np.where(
        df["Lucro Líquido (R$)"] > 0, 
        custos_fixos / df["Lucro Líquido (R$)"], 
        np.nan
    )

    if incluir_fixos:
        total_receita = df["Preco"].sum()
        if total_receita == 0:
            df["Custo Fixo Rateado (R$)"] = custos_fixos / max(len(df), 1)
        else:
            df["Custo Fixo Rateado (R$)"] = (df["Preco"] / total_receita) * custos_fixos

        df["Lucro Líquido (com fixos) (R$)"] = df["Lucro Líquido (R$)"] - df["Custo Fixo Rateado (R$)"]
        df["Margem Líquida (%)"] = np.where(
            df["Preco"] != 0, 
            (df["Lucro Líquido (com fixos) (R$)"] / df["Preco"]) * 100, 
            0
        )
        df["Preço Ideal c/ Fixos (R$)"] = np.where(
            1 - t - m > 0,
            (df["Custo"] + df["Outros Custos (R$)"] + df["Custo Fixo Rateado (R$)"]) / (1 - t - m),
            np.nan
        )
        df["Diferença Preço Ideal c/ Fixos (%)"] = np.where(
            df["Preco"] != 0,
            ((df["Preço Ideal c/ Fixos (R$)"] - df["Preco"]) / df["Preco"]) * 100,
            np.nan
        )

    return df.round(2)

# ----------------------------
# EXPORTAÇÃO
# ----------------------------
def exportar_excel(df_sem, df_com=None):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_sem.to_excel(writer, index=False, sheet_name="Resultados_Sem_Fixos")
        if df_com is not None:
            df_com.to_excel(writer, index=False, sheet_name="Resultados_Com_Fixos")
    return buffer.getvalue()

# ----------------------------
# MODELO EXCEL
# ----------------------------
def gerar_modelo_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Lucra+"
    ws.append(["Produto", "Custo", "Preco", "Taxa (%)", "Outros Custos (R$)"])
    ws.append(["Camiseta Azul", 25.0, 50.0, 2.5, 0.0])
    ws.append(["Caneca Logo", 18.0, 35.0, 3.0, 0.0])
    ws.append(["Bolo Pequeno", 12.0, 30.0, 5.0, 1.5])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col if c.value)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

# ----------------------------
# SIDEBAR CONFIG
# ----------------------------
st.sidebar.title("⚙️ Configurações")
margem_desejada = st.sidebar.number_input("Margem desejada (%)", 0.0, 99.0, 30.0)
custos_fixos = st.sidebar.number_input("Custos fixos mensais (R$)", 0.0, 1_000_000.0, 0.0, step=100.0)
incluir_fixos = st.sidebar.checkbox("Incluir custos fixos nos cálculos unitários", value=False)

# 👉 ADICIONAR OPÇÃO DE DASHBOARD APENAS PARA PREMIUM
opcoes_menu = ["📥 Importar / Adicionar", "📊 Resultados", "💾 Exportar", "ℹ️ Sobre"]
if plano_usuario == "Premium":
    opcoes_menu.insert(2, "📉 Dashboards")  # adiciona antes do Exportar

menu = st.sidebar.radio("📋 Navegação", opcoes_menu)


# -----------------------------
# RESULTADOS
# -----------------------------
if menu == "📊 Resultados":
    st.title("📊 Resultados e análises")

    if "dados" not in st.session_state or st.session_state.dados.empty:
        st.info("Nenhum produto cadastrado.")
    else:
        df_base = st.session_state.dados
        df_sem = calcular_resultados(df_base, margem_desejada, custos_fixos, incluir_fixos=False)
        df_com = calcular_resultados(df_base, margem_desejada, custos_fixos, incluir_fixos=True)
        df_full = df_com if incluir_fixos else df_sem

        lucro_col = "Lucro Líquido (com fixos) (R$)" if incluir_fixos and "Lucro Líquido (com fixos) (R$)" in df_full.columns else "Lucro Líquido (R$)"
        margem_col = "Margem Líquida (%)" if incluir_fixos and "Margem Líquída (%)" in df_full.columns else "Margem Atual (%)"

        ph1, ph2, ph3 = st.columns(3)
        place_lucro = ph1.empty()
        place_margem = ph2.empty()
        place_prod = ph3.empty()

        lucro_total_total = pd.to_numeric(df_full[lucro_col], errors="coerce").sum(min_count=1)
        margem_media_total = pd.to_numeric(df_full[margem_col], errors="coerce").mean()
        total_produtos_total = len(df_full)

        place_lucro.metric("💰 Lucro Total", f"R$ {lucro_total_total:.2f}")
        place_margem.metric("📉 Margem Média", f"{margem_media_total:.2f}%")
        place_prod.metric("📦 Produtos", total_produtos_total)

        st.markdown("---")
        st.subheader("📈 Detalhamento por produto (clique para selecionar)")

        gb = GridOptionsBuilder.from_dataframe(df_full)
        gb.configure_selection(selection_mode="multiple", use_checkbox=True)
        gb.configure_pagination(enabled=True, paginationAutoPageSize=False, paginationPageSize=10)
        gb.configure_default_column(resizable=True, sortable=True, filter=True, minWidth=160)
        gridOptions = gb.build()

        grid_response = AgGrid(
            df_full,
            gridOptions=gridOptions,
            update_mode=GridUpdateMode.SELECTION_CHANGED,
            fit_columns_on_grid_load=True,
            height=420,
            theme="alpine",
            enable_enterprise_modules=False
        )

        selected_raw = grid_response.get("selected_rows", [])
        if selected_raw is None:
            selected_records = []
        elif isinstance(selected_raw, pd.DataFrame):
            selected_records = selected_raw.to_dict("records")
        elif isinstance(selected_raw, dict):
            selected_records = [selected_raw]
        elif isinstance(selected_raw, list):
            selected_records = selected_raw
        else:
            try:
                selected_records = list(selected_raw)
            except Exception:
                selected_records = []

        if len(selected_records) > 0:
            df = pd.DataFrame(selected_records)
            for c in df.columns:
                if c != "Produto":
                    df[c] = pd.to_numeric(df[c], errors="coerce")
            selecionados = df["Produto"].astype(str).tolist()
            st.success(f"🔍 Filtro ativo: {', '.join(selecionados)}")
        else:
            df = df_full.copy()

        lucro_total = pd.to_numeric(df[lucro_col], errors="coerce").sum(min_count=1)
        margem_media = pd.to_numeric(df[margem_col], errors="coerce").mean()
        total_produtos = len(df)

        place_lucro.metric("💰 Lucro Total", f"R$ {lucro_total:.2f}")
        place_margem.metric("📉 Margem Média", f"{margem_media:.2f}%")
        place_prod.metric("📦 Produtos", total_produtos)

        st.markdown("---")
        st.markdown("### Gráfico de Margem por Produto")

        fig, ax = plt.subplots(figsize=(8, max(3, 0.25 * len(df))))
        ax.barh(
            df["Produto"], 
            df[margem_col], 
            color=["green" if x >= margem_desejada else "red" for x in df[margem_col]]
        )
        ax.set_xlabel(margem_col)
        ax.grid(axis="x", linestyle="--", alpha=0.5)
        st.pyplot(fig)

# ----------------------------
# IMPORTAR / ADICIONAR
# ----------------------------
elif menu == "📥 Importar / Adicionar":
    st.title("📥 Importar produtos ou adicionar manualmente")
    col1, col2 = st.columns(2)
    with col1:
        arquivo = st.file_uploader("Selecione o arquivo Excel (.xlsx ou .xls)", type=["xlsx", "xls"])

        if arquivo:
            df = pd.read_excel(arquivo)
            dados_atual = st.session_state.get("dados", pd.DataFrame())

            # --- BLOQUEIO PARA USUÁRIOS FREE ---
            total_novos = len(dados_atual) + len(df)
            if plano_usuario == "Free" and total_novos > 3:
                st.error("🚫 Usuários Free só podem ter **até 3 produtos cadastrados**.\n\nRemova alguns produtos ou faça upgrade para o Premium.")
                st.stop()

            st.session_state.dados = pd.concat([dados_atual, df], ignore_index=True)
            st.success(f"{len(df)} produtos importados com sucesso!")
            
        modelo_excel = gerar_modelo_excel()
        st.download_button("📘 Baixar modelo Excel (.xlsx)", modelo_excel, "Modelo_Lucra_Plus.xlsx")
    with col2:
        st.subheader("📝 Adicionar Produto Manualmente")
        with st.form("novo_produto", clear_on_submit=True):
            nome = st.text_input("Produto")
            custo = st.number_input("Custo (R$)", 0.0)
            preco = st.number_input("Preço (R$)", 0.0)
            taxa = st.number_input("Taxa (%)", 0.0)
            outros = st.number_input("Outros custos (R$)", 0.0)
            add = st.form_submit_button("Adicionar ➕")
     
                
            if add and nome:

                # --- BLOQUEIO PARA USUÁRIOS FREE ---
                dados_atual = st.session_state.get("dados", pd.DataFrame())
                if plano_usuario == "Free" and len(dados_atual) >= 3:
                    st.error("🚫 Usuários Free podem cadastrar no máximo **3 produtos**.\n\nFaça upgrade para o plano Premium para cadastrar ilimitado.")
                    st.stop()

                novo = pd.DataFrame([{
                    "Produto": nome,
                    "Custo": custo,
                    "Preco": preco,
                    "Taxa (%)": taxa,
                    "Outros Custos (R$)": outros
                }])
                st.session_state.dados = pd.concat([dados_atual, novo], ignore_index=True)
                st.success(f"Produto '{nome}' adicionado.")


# ----------------------------
# EXPORTAR
# ----------------------------
elif menu == "💾 Exportar":
    st.title("💾 Exportar resultados")
    if "dados" not in st.session_state or st.session_state.dados.empty:
        st.info("Nenhum produto cadastrado.")
    else:
        df_sem = calcular_resultados(st.session_state.dados, margem_desejada, custos_fixos, incluir_fixos=False)
        df_com = calcular_resultados(st.session_state.dados, margem_desejada, custos_fixos, incluir_fixos=True)
        excel = exportar_excel(df_sem, df_com)
        st.success("✅ Arquivo Excel gerado com abas de comparação.")
        st.download_button(
            "📊 Baixar Excel (.xlsx)", 
            excel, 
            f"Lucra_Resultados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

# ----------------------------
# DASHBOARDS (APENAS PREMIUM)
# ----------------------------
elif menu == "📉 Dashboards":
    if plano_usuario != "Premium":
        st.error("🚫 Esta funcionalidade é exclusiva para usuários Premium.")
        st.stop()

    st.title("📉 Dashboards (Premium)")

    if "dados" not in st.session_state or st.session_state.dados.empty:
        st.info("Nenhum produto cadastrado.")
        st.stop()

    df_base = st.session_state.dados.copy()

    # Calcula resultados
    df_sem = calcular_resultados(df_base, margem_desejada, custos_fixos, incluir_fixos=False)
    df_com = calcular_resultados(df_base, margem_desejada, custos_fixos, incluir_fixos=True)
    df = df_com if incluir_fixos else df_sem

    st.markdown("## 📌 Indicadores Simples")

    lucro_col = "Lucro Líquido (R$)" if not incluir_fixos else "Lucro Líquido (com fixos) (R$)"
    margem_col = "Margem Atual (%)" if not incluir_fixos else "Margem Líquida (%)"

    total_produtos = len(df)
    preco_medio = df["Preco"].mean()
    custo_medio = df["Custo"].mean()
    lucro_total = df[lucro_col].sum()
    margem_media = df[margem_col].mean()

    col1, col2, col3 = st.columns(3)
    col1.metric("🏷️ Preço Médio", f"R$ {preco_medio:.2f}")
    col2.metric("💵 Custo Médio", f"R$ {custo_medio:.2f}")    
    col3.metric("📉 Margem Média", f"{margem_media:.2f}%")


    st.markdown("---")
    st.markdown("## 📈 Gráficos")

    # ===============================
    # 1) TOP 5 PRODUTOS POR LUCRO
    # ===============================
    st.subheader("🏆 Top 5 Produtos por Lucro")

    df_top5 = df.sort_values(lucro_col, ascending=False).head(5)

    fig1, ax1 = plt.subplots(figsize=(8, 4))
    ax1.barh(df_top5["Produto"], df_top5[lucro_col], color="#4CAF50")
    ax1.invert_yaxis()
    ax1.set_xlabel("Lucro (R$)")
    ax1.grid(axis="x", linestyle="--", alpha=0.3)
    st.pyplot(fig1)


    # ===============================
    # 2) MARGEM POR PRODUTO (TOP 10)
    # ===============================
    st.subheader("📊 Margem por Produto (Top 10)")

    df_top_margem = df.sort_values(margem_col, ascending=False).head(10)

    fig3, ax3 = plt.subplots(figsize=(8, 4))
    ax3.barh(df_top_margem["Produto"], df_top_margem[margem_col], color="#2196F3")
    ax3.invert_yaxis()
    ax3.set_xlabel("Margem (%)")
    ax3.grid(axis="x", linestyle="--", alpha=0.3)
    st.pyplot(fig3)

# ----------------------------
# SOBRE
# ----------------------------
elif menu == "ℹ️ Sobre":
    st.title("💰 Sobre o Lucra+")

    st.markdown("""
    ### 💼 **Lucra+ v0.21**
    O **Lucra+** é uma aplicação desenvolvida para ajudar **empreendedores e gestores** 
    a entender e otimizar a **margem de lucro dos seus produtos**, com base em custos,
    taxas e metas de rentabilidade.

    ---
    #### ⚙️ **Como o Lucra+ funciona**
    1. Você importa ou cadastra seus produtos com preço, custo e taxas.  
    2. O sistema calcula automaticamente:
       - Lucro líquido (com e sem custos fixos)  
       - Margem atual e ideal  
       - Preço ideal para atingir a margem desejada  
       - Ponto de equilíbrio  
    3. Você visualiza os resultados em tabelas, gráficos e indicadores dinâmicos.

    ---
    #### 🚀 **Principais recursos**
    - Upload de planilhas Excel (.xlsx / .xls)  
    - Adição manual de produtos  
    - Cálculos automáticos de lucro e margem  
    - Filtro de produtos com atualização imediata  
    - Exportação dos resultados para Excel  
    - Gráfico visual de desempenho por produto  

    ---
    #### 🧩 **Tecnologias utilizadas**
    - Streamlit  
    - Pandas  
    - Matplotlib  
    - OpenPyXL  

    ---
    #### 💬 **Agradecimento**
    Este projeto foi criado com o objetivo de **tornar o controle de margens simples e acessível**.

    ---
    🏷️ **Versão:** 0.21  
    📅 **Última atualização:** Novembro/2025  
    """)
